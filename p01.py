import csv
import sys
import datetime
import sqlite3
import fnmatch
import openpyxl

from PyQt5 import QtCore
from PyQt5.QtWidgets import QApplication, QMainWindow, QFileDialog, QTableWidgetItem, QMessageBox, QDialog, \
    QColorDialog
from PyQt5.QtGui import QPixmap, QIcon, QFont, QColor, QFontDatabase
from PyQt5.QtCore import Qt, QDate, QPropertyAnimation, QSize

from UI.Main_window import Ui_MainWindow
from UI.Table import Ui_table
from UI.Add_Applicant import Ui_add_student
from UI.Viewing_Applicants import Ui_viewing_applicants
from UI.Print_window import Ui_print
from UI.About_programm import Ui_about_programm
from UI.Dialog import Ui_Dialog


def except_hook(cls, exception, traceback):
    sys.__excepthook__(cls, exception, traceback)


# главное меню
class Main_window(Ui_MainWindow, QMainWindow):
    def __init__(self):
        super().__init__()
        self.setupUi(self)
        self.main_window()

        self.file_1 = ""
        self.file_2 = ""
        self.file_3 = ""
        self.file_4 = ""
        self.file_5 = ""
        self.file_6 = ""
        self.file_7 = ""

        # установка надписи в главном меню
        current_hour = (str(datetime.datetime.now()).split())[1].split(":")[0]
        if current_hour[0] == "0":
            current_hour = int(current_hour[1])

        else:
            current_hour = int(current_hour)

        if 3 < current_hour < 12:
            self.current_time_of_day.setText("Доброе утро!")

        elif 12 <= current_hour < 17:
            self.current_time_of_day.setText("Добрый день!")

        elif 17 <= current_hour < 22:
            self.current_time_of_day.setText("Добрый вечер!")

        else:
            self.current_time_of_day.setText("Доброй ночи!")
        self.setWindowIcon(QIcon("images/icon.png"))

        with open("sys_files/recent_files.txt") as file:
            file_spisok = [str(i) for i in file]

            if len(file_spisok) > 0:
                self.empty_file_label.hide()
                for i in range(len(file_spisok)):
                    if i == 0:
                        self.recent_file_1.setText(file_spisok[0].split("/")[-1])
                        self.file_1 = file_spisok[0]

                    elif i == 1:
                        self.recent_file_2.setText(file_spisok[1].split("/")[-1])
                        self.file_2 = file_spisok[1]
                        self.recent_file_2.show()

                    elif i == 2:
                        self.recent_file_3.setText(file_spisok[2].split("/")[-1])
                        self.file_3 = file_spisok[2]
                        self.recent_file_3.show()

                    elif i == 3:
                        self.recent_file_4.setText(file_spisok[3].split("/")[-1])
                        self.file_4 = file_spisok[3]
                        self.recent_file_4.show()

                    elif i == 4:
                        self.recent_file_5.setText(file_spisok[4].split("/")[-1])
                        self.file_5 = file_spisok[4]
                        self.recent_file_5.show()

                    elif i == 5:
                        self.recent_file_6.setText(file_spisok[5].split("/")[-1])
                        self.file_6 = file_spisok[5]
                        self.recent_file_6.show()

                    elif i == 6:
                        self.recent_file_7.setText(file_spisok[6].split("/")[-1])
                        self.file_7 = file_spisok[6]
                        self.recent_file_7.show()

    def main_window(self):  # главное меню программы
        # обработка нажатий на кнопки
        self.open_table.clicked.connect(self.open_table_function)
        self.open_file_from_main.clicked.connect(self.open_file)
        self.add_student.clicked.connect(self.open_add_student)
        self.open_list_of_applicants.clicked.connect(self.open_viewing_applicants)
        self.print_documents.clicked.connect(self.open_print_documents)
        self.parameters.clicked.connect(self.open_about_programm)
        self.exit.clicked.connect(self.exit_p)

        # обработка нажатий на кнопки с недавними файлами

        self.recent_file_1.clicked.connect(self.open_file_from_recent)
        self.recent_file_2.clicked.connect(self.open_file_from_recent)
        self.recent_file_3.clicked.connect(self.open_file_from_recent)
        self.recent_file_4.clicked.connect(self.open_file_from_recent)
        self.recent_file_5.clicked.connect(self.open_file_from_recent)
        self.recent_file_6.clicked.connect(self.open_file_from_recent)
        self.recent_file_7.clicked.connect(self.open_file_from_recent)

    def open_file_from_recent(self):
        if self.sender() == self.recent_file_1:
            self.file = self.file_1.replace("\n", "")

        elif self.sender() == self.recent_file_2:
            self.file = self.file_2.replace("\n", "")

        elif self.sender() == self.recent_file_3:
            self.file = self.file_3.replace("\n", "")

        elif self.sender() == self.recent_file_4:
            self.file = self.file_4.replace("\n", "")

        elif self.sender() == self.recent_file_5:
            self.file = self.file_5.replace("\n", "")

        elif self.sender() == self.recent_file_6:
            self.file = self.file_6.replace("\n", "")

        elif self.sender() == self.recent_file_7:
            self.file = self.file_7.replace("\n", "")

        self.table = Window()
        self.table.show()
        Main_window.close(self)

        if fnmatch.fnmatch(self.file, "*.csv"):
            dialog = Csv_encoding()
            encode = ""

            if dialog.exec_():
                if dialog.encoding.currentText() == "UTF-8 (по умолчанию)":
                    encode = "utf8"

                elif dialog.encoding.currentText() == "ASCII":
                    encode = "ASCII"

                elif dialog.encoding.currentText() == "Windows-1251 (для кириллических алфавитов)":
                    encode = "Windows-1251"

                if encode != "":
                    with open(self.file, encoding=encode, errors="replace") as csvfile:
                        if dialog.separator.currentText() == "Точка с запятой":
                            reader = csv.reader(csvfile, delimiter=";")  # открываю файл

                        elif dialog.separator.currentText() == "Табуляция":
                            reader = csv.reader(csvfile, delimiter="\t")

                        elif dialog.separator.currentText() == "Пробел":
                            reader = csv.reader(csvfile, delimiter=" ")

                        elif dialog.separator.currentText() == "Запятая":
                            reader = csv.reader(csvfile, delimiter=",")

                        elif dialog.separator.currentText() == "Двоеточие":
                            reader = csv.reader(csvfile, delimiter=":")

                        for i, row in enumerate(reader):  # заполняю таблицу данными
                            self.table.table_main.setRowCount(self.table.table_main.rowCount() + 1)

                            for j, element in enumerate(row):
                                self.table.table_main.setItem(i, j, QTableWidgetItem(element))

                        if len(self.file) > 0:
                            # далее буду открывать таблицу с файлом
                            self.setWindowTitle(self.file.split("/")[-1])

        elif fnmatch.fnmatch(self.file, "*.xls") or fnmatch.fnmatch(self.file, "*.xlsx"):
            wb = openpyxl.load_workbook(self.file)
            name = wb.sheetnames[0]
            table = wb[name]
            data = list()

            for row in table:
                spisok = list()
                for elem in row:
                    spisok.append(elem.value)
                data.append(spisok)

            for row in range(len(data)):
                self.table.table_main.setRowCount(self.table.table_main.rowCount() + 1)
                for column in range(len(data[row])):
                    self.table.table_main.setColumnCount(max(self.table.table_main.columnCount(), len(data[row])))
                    self.table.table_main.setItem(row, column, QTableWidgetItem(str(data[row][column])))

        elif fnmatch.fnmatch(self.file, "*.sqlite"):
            pass

    # выход из программы
    def exit_p(self):
        self.close()

    def open_print_documents(self):
        self.doc = Print_documents()
        self.doc.show()
        self.close()

    def open_viewing_applicants(self):
        self.viewing = Viewing_applicants()
        self.viewing.show()
        self.close()

    # отображение окна "О программе"
    def open_about_programm(self):
        self.information = About()
        self.information.show()

    def open_file(self):  # открытие файла через главное меню
        self.file = QFileDialog.getOpenFileName(self, "Выберите файл", "/",
                                                "CSV-таблица (*.csv);; Таблица (*.xls, *.xlsx);; База данных (*.sqlite)")[
            0]

        if len(self.file) > 0:
            # далее буду открывать таблицу с файлом
            self.table = Window()
            self.table.show()
            Main_window.close(self)

            write_permission = True
            with open("recent_files.txt") as file1:
                for row in file1:
                    if self.file in row:
                        write_permission = False
                        break

            if write_permission:
                with open("recent_files.txt", "a") as file2:
                    file2.write(f"{self.file}\n")

            if fnmatch.fnmatch(self.file, "*.csv"):
                dialog = Csv_encoding()
                encode = ""

                if dialog.exec_():
                    if dialog.encoding.currentText() == "UTF-8 (по умолчанию)":
                        encode = "utf8"

                    elif dialog.encoding.currentText() == "ASCII":
                        encode = "ASCII"

                    elif dialog.encoding.currentText() == "Windows-1251 (для кириллических алфавитов)":
                        encode = "Windows-1251"

                    if encode != "":
                        with open(self.file, encoding=encode, errors="replace") as csvfile:
                            if dialog.separator.currentText() == "Точка с запятой":
                                reader = csv.reader(csvfile, delimiter=";")  # открываю файл

                            elif dialog.separator.currentText() == "Табуляция":
                                reader = csv.reader(csvfile, delimiter="\t")

                            elif dialog.separator.currentText() == "Пробел":
                                reader = csv.reader(csvfile, delimiter=" ")

                            elif dialog.separator.currentText() == "Запятая":
                                reader = csv.reader(csvfile, delimiter=",")

                            elif dialog.separator.currentText() == "Двоеточие":
                                reader = csv.reader(csvfile, delimiter=":")

                            for i, row in enumerate(reader):  # заполняю таблицу данными
                                self.table.table_main.setRowCount(self.table.table_main.rowCount() + 1)

                                for j, element in enumerate(row):
                                    self.table.table_main.setItem(i, j, QTableWidgetItem(element))

                            if len(self.file) > 0:
                                # далее буду открывать таблицу с файлом
                                self.setWindowTitle(self.file.split("/")[-1])

            elif fnmatch.fnmatch(self.file, "*.xls") or fnmatch.fnmatch(self.file, "*.xlsx"):
                wb = openpyxl.load_workbook(self.file)
                name = wb.sheetnames[0]
                table = wb[name]
                data = list()

                for row in table:
                    spisok = list()
                    for elem in row:
                        spisok.append(elem.value)
                    data.append(spisok)

                self.table.table_main.setRowCount(0)
                self.table.table_main.setColumnCount(0)

                for row in range(len(data)):
                    self.table.table_main.setRowCount(self.table.table_main.rowCount() + 1)
                    for column in range(len(data[row])):
                        self.table.table_main.setColumnCount(max(self.table.table_main.columnCount(), len(data[row])))
                        self.table.table_main.setItem(row, column, QTableWidgetItem(str(data[row][column])))

            elif fnmatch.fnmatch(self.file, "*.sqlite"):
                pass

    # открытие новой таблицы
    def open_table_function(self):
        self.close()
        self.table = Window()
        self.table.show()

    def open_add_student(self):
        self.student = Add_student()
        self.student.show()
        self.close()


# окно с таблицей
class Window(Ui_table, QMainWindow):
    def __init__(self):
        super(Window, self).__init__()
        self.setupUi(self)
        self.new_table()
        self.last_column_header = ""
        self.setWindowIcon(QIcon("images/icon.png"))

        self.file = ""
        self.encode = ""
        self.separator = ""

        self.color_rgv = (255, 190, 111)
        self.flag = True
        self.cells_color = list()

    def hex_to_rgb(self, color_hex):
        color_hex = color_hex.lstrip("#")
        return tuple([int(color_hex[i:i + 2], 16) for i in range(0, 5, 2)])

    # подтверждение выхода
    def closeEvent(self, e):
        if "*" in self.windowTitle():
            confirm = QMessageBox.question(self, "Подтверждение закрытия окна",
                                           "Внимание! Все несохранённые изменения будут потеряны",
                                           QMessageBox.Yes | QMessageBox.No)
            if confirm == QMessageBox.Yes:
                self.main_m = Main_window()
                self.main_m.show()
                e.accept()

            else:
                e.ignore()
        else:
            self.main_m = Main_window()
            self.main_m.show()
            e.accept()

    def keyPressEvent(self, event):
        if event.key() == Qt.Key_Delete:
            self.clear_values()

    def confirm_exit(self):
        self.close()

    def new_table(self):
        self.search_line.setPlaceholderText("Поиск по таблице")
        self.table_main.itemChanged.connect(self.add_star)
        # выход в главное меню
        self.action_2.triggered.connect(self.confirm_exit)
        self.action_2.setShortcut("Ctrl+E")

        # открытие файла
        self.open_file.triggered.connect(self.open_file_in_table)
        self.open_file.setShortcut("Ctrl+O")

        # сохранение файла
        self.save_file.triggered.connect(self.save_file_function)
        self.save_file.setShortcut("Ctrl+S")
        self.save_button.clicked.connect(self.save_file_function)

        self.save_as.triggered.connect(self.save_file_as_function)

        self.setWindowTitle("Безымянный")

        alphabit = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
        spisok = list()

        for i in range(50):
            if i > 0:
                spisok.append(alphabit[i % 26] * ((i // 26) + 1))

            else:
                spisok.append("А")

        self.last_column_header = spisok[-1]
        self.table_main.setHorizontalHeaderLabels(spisok)
        self.table_main.currentCellChanged.connect(self.set_current_cell)
        self.search_line.textChanged.connect(self.search_in_table)

        # редактирование текста в ячейке
        self.bold_text.clicked.connect(self.set_bold)
        self.italic_text.clicked.connect(self.set_italic)
        self.underlined_text.clicked.connect(self.set_underlined)

        self.fontComboBox.currentFontChanged.connect(self.change_font)
        self.text_size.currentTextChanged.connect(self.change_text_size)
        self.text_color.currentTextChanged.connect(self.change_color_text)

        self.align_left.clicked.connect(self.align_left_function)
        self.align_right.clicked.connect(self.align_right_function)
        self.align_center.clicked.connect(self.align_centre_function)

        # перекраска ячеек
        self.white_color.clicked.connect(self.repaint_cell_white)
        self.red_color.clicked.connect(self.repaint_cell_red)
        self.green_color.clicked.connect(self.repaint_cell_green)
        self.orange_color.clicked.connect(self.repaint_cell_other)
        self.add_new_color.clicked.connect(self.set_new_color_cell)

    def search_in_table(self):
        items = self.table_main.findItems(self.search_line.text(), Qt.MatchContains)
        spisok = [(i.row(), i.column()) for i in items if i is not None]
        all_items = []
        # доработать
        for i in range(self.table_main.rowCount()):
            for j in range(self.table_main.columnCount()):
                if self.table_main.item(i, j) is not None:
                    all_items.append((i, j, self.table_main.item(i, j).background().color()))

        if self.search_line.text() != "":
            for i in range(len(all_items)):
                self.table_main.item(all_items[i][0], all_items[i][1]).setBackground(QColor(all_items[i][2]))

            for i in items:
                if i is not None:
                    self.table_main.item(i.row(), i.column()).setBackground(QColor("#EEF019"))

        else:
            for i in range(len(all_items)):
                self.table_main.item(all_items[i][0], all_items[i][1]).setBackground(QColor(all_items[i][2]))

    # перекраска цветов
    def repaint_cell_white(self):
        spisok = [(i.row(), i.column()) for i in self.table_main.selectedItems()]

        for row, column in spisok:
            self.table_main.item(row, column).setBackground(QColor(255, 255, 255))

    def repaint_cell_red(self):
        spisok = [(i.row(), i.column()) for i in self.table_main.selectedItems()]

        for row, column in spisok:
            self.table_main.item(row, column).setBackground(QColor(246, 97, 81))

    def repaint_cell_green(self):
        spisok = [(i.row(), i.column()) for i in self.table_main.selectedItems()]

        for row, column in spisok:
            self.table_main.item(row, column).setBackground(QColor(143, 240, 164))

    def repaint_cell_other(self):
        spisok = [(i.row(), i.column()) for i in self.table_main.selectedItems()]

        for row, column in spisok:
            self.table_main.item(row, column).setBackground(
                QColor(self.color_rgv[0], self.color_rgv[1], self.color_rgv[2]))

    # замена оранжевого цвета на другой
    def set_new_color_cell(self):
        color = QColorDialog.getColor()
        color = self.hex_to_rgb(color.name())
        self.color_rgv = color

        color = f"background-color: rgb{color};\n"
        color_style = ["QPushButton{\n", "border:  none;\n",
                       "margin:  0px;\n",
                       "padding: 0px;\n", "border-radius: 10px;",
                       color,
                       "}"]

        self.orange_color.setStyleSheet("".join(color_style))
        self.orange_color.setText("Другой")

    # установка стилей текста
    def change_color_text(self):
        pass

    def align_right_function(self):
        spisok = [(i.row(), i.column()) for i in self.table_main.selectedItems()]

        for row, column in spisok:
            font = self.table_main.item(row, column).font()
            font.setPointSize(self.table_main.item(row, column).font().pointSize())
            back = self.table_main.item(row, column).background()

            item = QTableWidgetItem(self.table_main.item(row, column).text())
            item.setTextAlignment(QtCore.Qt.AlignRight | QtCore.Qt.AlignVCenter)

            if self.table_main.item(row, column).font().italic():
                font.setItalic(True)

            if self.table_main.item(row, column).font().underline():
                font.setUnderline(True)

            if self.table_main.item(row, column).font().bold():
                font.setBold(True)

            self.table_main.setItem(row, column, item)
            self.table_main.item(row, column).setFont(font)
            self.table_main.item(row, column).setBackground(back)

    def align_centre_function(self):
        spisok = [(i.row(), i.column()) for i in self.table_main.selectedItems()]

        for row, column in spisok:
            font = self.table_main.item(row, column).font()
            font.setPointSize(self.table_main.item(row, column).font().pointSize())
            back = self.table_main.item(row, column).background()

            item = QTableWidgetItem(self.table_main.item(row, column).text())
            item.setTextAlignment(QtCore.Qt.AlignVCenter | QtCore.Qt.AlignHCenter)

            if self.table_main.item(row, column).font().italic():
                font.setItalic(True)

            if self.table_main.item(row, column).font().underline():
                font.setUnderline(True)

            if self.table_main.item(row, column).font().bold():
                font.setBold(True)

            self.table_main.setItem(row, column, item)
            self.table_main.item(row, column).setFont(font)
            self.table_main.item(row, column).setBackground(back)

    def align_left_function(self):
        spisok = [(i.row(), i.column()) for i in self.table_main.selectedItems()]

        for row, column in spisok:
            font = self.table_main.item(row, column).font()
            font.setPointSize(self.table_main.item(row, column).font().pointSize())
            back = self.table_main.item(row, column).background()

            if self.table_main.item(row, column).font().italic():
                font.setItalic(True)

            if self.table_main.item(row, column).font().underline():
                font.setUnderline(True)

            if self.table_main.item(row, column).font().bold():
                font.setBold(True)

            item = QTableWidgetItem(self.table_main.item(row, column).text())
            item.setTextAlignment(QtCore.Qt.AlignLeft | QtCore.Qt.AlignVCenter)

            self.table_main.setItem(row, column, item)
            self.table_main.item(row, column).setFont(font)
            self.table_main.item(row, column).setBackground(back)

    def change_font(self):
        spisok = [(i.row(), i.column()) for i in self.table_main.selectedItems()]

        for row, column in spisok:
            font = QFont(self.fontComboBox.currentFont())
            font.setPointSize(self.table_main.item(row, column).font().pointSize())
            back = self.table_main.item(row, column).background()

            if self.table_main.item(row, column) is not None:
                if self.table_main.item(row, column).font().italic():
                    font.setItalic(True)

                if self.table_main.item(row, column).font().underline():
                    font.setUnderline(True)

                if self.table_main.item(row, column).font().bold():
                    font.setBold(True)

                self.table_main.item(row, column).setFont(font)
                self.table_main.item(row, column).setBackground(back)

    def change_text_size(self):
        if self.flag:
            spisok = [(i.row(), i.column()) for i in self.table_main.selectedItems()]

            for row, column in spisok:
                font = self.table_main.item(row, column).font()
                font.setPointSize(int(self.text_size.currentText()))

                if self.table_main.item(row, column) is not None:
                    if self.table_main.item(row, column).font().italic():
                        font.setItalic(True)

                    if self.table_main.item(row, column).font().underline():
                        font.setUnderline(True)

                    if self.table_main.item(row, column).font().bold():
                        font.setBold(True)

                    self.table_main.item(row, column).setFont(font)
        else:
            self.flag = True

    def set_bold(self):
        spisok = [(i.row(), i.column()) for i in self.table_main.selectedItems()]

        for row, column in spisok:
            font = self.table_main.item(row, column).font()
            font.setPointSize(int(self.text_size.currentText()))
            font.setBold(True)

            if self.table_main.item(row, column) is not None:
                if self.table_main.item(row, column).font().italic():
                    font.setItalic(True)

                if self.table_main.item(row, column).font().underline():
                    font.setUnderline(True)

                if self.table_main.item(row, column).font().bold():
                    font.setBold(False)

                self.table_main.item(row, column).setFont(font)

    def set_italic(self):
        spisok = [(i.row(), i.column()) for i in self.table_main.selectedItems()]

        for row, column in spisok:
            font = self.table_main.item(row, column).font()
            font.setPointSize(int(self.text_size.currentText()))
            font.setItalic(True)

            if self.table_main.item(row, column) is not None:
                if self.table_main.item(row, column).font().italic():
                    font.setItalic(False)

                if self.table_main.item(row, column).font().underline():
                    font.setUnderline(True)

                if self.table_main.item(row, column).font().bold():
                    font.setBold(True)

                self.table_main.item(row, column).setFont(font)

    def set_underlined(self):
        spisok = [(i.row(), i.column()) for i in self.table_main.selectedItems()]

        for row, column in spisok:
            font = self.table_main.item(row, column).font()
            font.setPointSize(int(self.text_size.currentText()))
            font.setUnderline(True)

            if self.table_main.item(row, column) is not None:
                if self.table_main.item(row, column).font().italic():
                    font.setItalic(True)

                if self.table_main.item(row, column).font().underline():
                    font.setUnderline(False)

                if self.table_main.item(row, column).font().bold():
                    font.setBold(True)

                self.table_main.item(row, column).setFont(font)

    def set_current_cell(self):
        alphabit = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
        spisok = sorted([(i.row(), i.column()) for i in self.table_main.selectedIndexes()])

        if len(spisok) > 1:
            self.current_cell.setText(
                f"{alphabit[spisok[0][1] % 26] * (spisok[0][1] // 26 + 1)}{spisok[0][0] + 1}:\
{alphabit[spisok[-1][1] % 26] * (spisok[-1][1] // 26 + 1)}\
{spisok[-1][0] + 1}")

        else:
            self.current_cell.setText(
                f"{alphabit[self.table_main.currentColumn() % 26] * (int(self.table_main.currentColumn()) // 26 + 1)}\
{self.table_main.currentRow() + 1}")

        if self.table_main.currentItem() is not None:
            self.flag = False
            self.text_size.setCurrentText(str(self.table_main.currentItem().font().pointSize()))

        else:
            self.flag = False
            self.text_size.setCurrentText("11")
        self.flag = True

    # добавляет * для того, чтобы понять что были сделаны изменения
    def add_star(self):
        if self.table_main.currentItem() is not None:
            if self.table_main.currentItem().text() != "":
                if self.windowTitle().count("*") == 0:
                    title = self.windowTitle() + "*"
                    self.setWindowTitle(title)

    # очистка значений в ячейках
    def clear_values(self):
        spisok = [(i.row(), i.column()) for i in self.table_main.selectedItems()]
        for row, column in spisok:
            self.table_main.item(row, column).setText("")

    # открытие файла прямо в таблице
    def open_file_in_table(self):
        self.file = QFileDialog.getOpenFileName(self, "Выберите файл", "/",
                                                "CSV-таблица (*.csv);; Таблица (*.xls, *.xlsx);; База данных (*.sqlite)"
                                                )[0]

        if fnmatch.fnmatch(self.file, "*.csv"):
            dialog = Csv_encoding()
            self.encode = ""

            if dialog.exec_():
                if dialog.encoding.currentText() == "UTF-8 (по умолчанию)":
                    self.encode = "utf8"

                elif dialog.encoding.currentText() == "ASCII":
                    self.encode = "ASCII"

                elif dialog.encoding.currentText() == "Windows-1251 (для кириллических алфавитов)":
                    self.encode = "Windows-1251"

                if self.encode != "":
                    write_permission = True
                    with open("recent_files.txt") as file1:
                        for row in file1:
                            if self.file in row:
                                write_permission = False
                                break

                    if write_permission:
                        with open("recent_files.txt", "a") as file2:
                            file2.write(f"{self.file}\n")

                    with open(self.file, encoding=self.encode, errors="replace") as csvfile:
                        if dialog.separator.currentText() == "Точка с запятой":
                            self.separator = ";"
                            reader = csv.reader(csvfile, delimiter=";")  # открываю файл

                        elif dialog.separator.currentText() == "Табуляция":
                            self.separator = "\t"
                            reader = csv.reader(csvfile, delimiter="\t")

                        elif dialog.separator.currentText() == "Пробел":
                            self.separator = " "
                            reader = csv.reader(csvfile, delimiter=" ")

                        elif dialog.separator.currentText() == "Запятая":
                            self.separator = ","
                            reader = csv.reader(csvfile, delimiter=",")

                        elif dialog.separator.currentText() == "Двоеточие":
                            self.separator = ":"
                            reader = csv.reader(csvfile, delimiter=":")

                        self.table_main.setRowCount(0)
                        self.table_main.setColumnCount(0)

                        for i, row in enumerate(reader):  # заполняю таблицу данными
                            self.table_main.setRowCount(self.table_main.rowCount() + 1)
                            self.table_main.setColumnCount(max(self.table_main.columnCount(), len(row)))
                            for j, element in enumerate(row):
                                if len(row) != 0:
                                    self.table_main.setItem(i, j, QTableWidgetItem(element))

        elif fnmatch.fnmatch(self.file, "*.xls") or fnmatch.fnmatch(self.file, "*.xlsx"):
            wb = openpyxl.load_workbook(self.file)
            name = wb.sheetnames[0]
            table = wb[name]
            data = list()

            write_permission = True
            with open("recent_files.txt") as file1:
                for row in file1:
                    if self.file in row:
                        write_permission = False
                        break

            if write_permission:
                with open("recent_files.txt", "a") as file2:
                    file2.write(f"{self.file}\n")

            for row in table:
                spisok = list()
                for elem in row:
                    spisok.append(elem.value)
                data.append(spisok)

            self.table_main.setRowCount(0)
            self.table_main.setColumnCount(0)

            for row in range(len(data)):
                self.table_main.setRowCount(self.table_main.rowCount() + 1)
                for column in range(len(data[row])):
                    self.table_main.setColumnCount(max(self.table_main.columnCount(), len(data[row])))
                    self.table_main.setItem(row, column, QTableWidgetItem(str(data[row][column])))

        elif fnmatch.fnmatch(self.file, "*.sqlite"):
            pass

        if len(self.file) > 0:
            # далее буду открывать таблицу с файлом
            self.setWindowTitle(self.file.split("/")[-1])

        alphabit = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
        spisok = list()

        for i in range(50):
            if i > 0:
                spisok.append(alphabit[i % 26] * ((i // 26) + 1))

            else:
                spisok.append("А")

        self.table_main.setHorizontalHeaderLabels(spisok)

    def save_file_function(self):
        if self.windowTitle() != "Безымянный" and self.windowTitle() != "Безымянный*":
            with open(self.file, mode="w", encoding=self.encode) as csvfile:
                writer = csv.writer(csvfile, delimiter=self.separator)

                for i in range(self.table_main.rowCount()):
                    row = []

                    for j in range(self.table_main.columnCount()):
                        item = self.table_main.item(i, j)

                        if item is not None:
                            row.append(item.text())

                    writer.writerow(row)

                write_permission = True
                with open("recent_files.txt") as file1:
                    for row in file1:
                        if self.file in row:
                            write_permission = False
                            break

                if write_permission:
                    with open("recent_files.txt", "a") as file2:
                        file2.write(f"{self.file}\n")

            if "*" in self.windowTitle():
                title = self.windowTitle()[:-1]
                self.setWindowTitle(title)

        else:
            self.save_file_as_function()

    def save_file_as_function(self):
        file = QFileDialog.getSaveFileName(self, "Сохранить файл", "/",
                                           "CSV-файл, разделённый запятой(*.csv);;Таблица(*xls, *xlsx")[0]
        name_file = file.split("/")[-1]

        if ".csv" not in file:
            file += ".csv"

        with open(file, mode="w", encoding="utf8") as csvfile:
            writer = csv.writer(csvfile, delimiter=",")
            for i in range(self.table_main.rowCount()):
                row = []

                for j in range(self.table_main.columnCount()):
                    item = self.table_main.item(i, j)

                    if item is not None:
                        row.append(item.text())

                writer.writerow(row)

            write_permission = True
            with open("recent_files.txt") as file1:
                for row in file1:
                    if self.file in row:
                        write_permission = False
                        break

            if write_permission:
                with open("recent_files.txt", "a") as file2:
                    file2.write(f"{self.file}\n")

            if "*" in self.windowTitle() and name_file != "":
                title = self.windowTitle()[:-1]
                self.setWindowTitle(title)


class Csv_encoding(Ui_Dialog, QDialog):
    def __init__(self):
        super(Ui_Dialog, self).__init__()
        self.setupUi(self)
        self.setWindowIcon(QIcon("images/icon.png"))


# окно "Добавить абитуриента"
class Add_student(Ui_add_student, QMainWindow):
    def __init__(self):
        super(Ui_add_student, self).__init__()
        self.setupUi(self)
        self.proccess()
        self.setWindowIcon(QIcon("images/icon.png"))
        self.flag2 = True
        self.flag = True
        self.encode = "" \
                      ""
        self.hide_dialog = False

        # списки, содержащие ссылки на изображения для image labels
        self.oms_images = list()
        self.passport_images = list()
        self.snils_images = list()
        self.birth_certificate_images = list()
        self.achievements_images = list()
        self.certificate_images = list()

        self.move_widgets_in_first_column()
        self.move_widgets_in_second_column()
        self.move_widgets_in_second_column_2()
        self.move_widgets_in_second_column_3()

        self.open_and_close_calendar_issue()

    # подтверждение выхода
    def closeEvent(self, e):
        if not self.hide_dialog:
            confirm = QMessageBox.question(self, "Подтверждение закрытия окна",
                                           "Внимание! Все несохранённые изменения будут потеряны",
                                           QMessageBox.Yes | QMessageBox.No)
            if confirm == QMessageBox.Yes:
                self.window = Main_window()
                self.window.show()
                e.accept()

            else:
                e.ignore()
        else:
            self.hide_dialog = False

    def keyPressEvent(self, event):
        if event.key() == 16777220:
            self.focusWidget().clearFocus()

    def proccess(self):
        self.save.clicked.connect(self.check_data)
        self.cancel.clicked.connect(self.confirm_exit)

        # триггер календарей
        self.calendar_birthday.selectionChanged.connect(self.change_date)
        self.calendar_issue.selectionChanged.connect(self.change_date)

        # триггер при нажатии на лупу у кнопок
        self.open_explorer_passport.clicked.connect(self.set_image)
        self.open_explorer_snils.clicked.connect(self.set_image)
        self.open_explorer_achievements.clicked.connect(self.set_image)
        self.open_explorer_oms.clicked.connect(self.set_image)
        self.open_explorer_birth_certificate.clicked.connect(self.set_image)
        self.open_explorer_certificate.clicked.connect(self.set_image)
        self.open_explorer_passport_2.clicked.connect(self.set_image)

        # триггер открытия и закрытия календарей
        self.open_calendar_issue.clicked.connect(self.open_and_close_calendar_issue)
        self.open_calendar_birthday.clicked.connect(self.open_and_hide_calendar_birthday)

        self.open_path.clicked.connect(self.get_path)

        self.open_attachments_passport.clicked.connect(self.move_widgets_in_first_column)
        self.open_attachments_certificate_number.clicked.connect(self.move_attachments_certificate)
        self.open_attachments_snils.clicked.connect(self.move_widgets_in_second_column)
        self.open_attachments_oms.clicked.connect(self.move_widgets_in_second_column_2)
        self.open_attachments_birth_certificate.clicked.connect(self.move_widgets_in_second_column_3)

        self.about_applicant_textedit.textChanged.connect(self.length_text_edit)
        self.open_and_close_button.clicked.connect(self.close_and_open_side_panel)

        # очистка надписей, сообщающих об ошибке при изменении поля
        self.date_of_birth.dateChanged.connect(self.erase_the_inscription)
        self.date_of_issue.dateChanged.connect(self.erase_the_inscription)

        self.fio.textEdited.connect(self.erase_the_inscription)
        self.address.textEdited.connect(self.erase_the_inscription)
        self.path.textEdited.connect(self.erase_the_inscription)
        self.series_and_number_passport.textEdited.connect(self.erase_the_inscription)
        self.issued_by_whom.textEdited.connect(self.erase_the_inscription)
        self.snils.textEdited.connect(self.erase_the_inscription)
        self.certificate_number.textEdited.connect(self.erase_the_inscription)

        # при изменении региона
        self.region.currentTextChanged.connect(self.change_cities_and_villages)

    def set_image(self):
        supported_formats = ["png", "jpg", "jpeg", "ico", "webp", "svg"]

        try:
            file = \
                QFileDialog.getOpenFileName(self, "Выберите файл", "/", "Изображение (*.jpg, *.jpeg, *.png)")[0]

            if len(file) == 0:
                raise FileNotFoundError

            elif file.split(".")[-1] not in supported_formats:
                raise ValueError

            else:
                if self.sender() == self.open_explorer_passport:
                    if file not in self.passport_images:
                        self.passport_image.setPixmap(QPixmap(file))
                        self.passport_images.append(file)

                    else:
                        QMessageBox.warning(self, "Добавление изображения",
                                            "Данный файл уже загружен", QMessageBox.Ok)

                    self.left_button_passport.setHidden(False)
                    self.right_button_passport.setHidden(False)
                    self.delete_image_passport.setHidden(False)

                elif self.sender() == self.open_explorer_oms:
                    if file not in self.oms_images:
                        self.oms_image.setPixmap(QPixmap(file))
                        self.oms_images.append(file)

                    else:
                        QMessageBox.warning(self, "Добавление изображения",
                                            "Данный файл уже загружен", QMessageBox.Ok)

                    self.left_button_oms.setHidden(False)
                    self.right_button_oms.setHidden(False)
                    self.delete_image_oms.setHidden(False)

                elif self.sender() == self.open_explorer_birth_certificate:
                    if file not in self.birth_certificate_images:
                        self.birth_certificate_image.setPixmap(QPixmap(file))
                        self.birth_certificate_images.append(file)

                    else:
                        QMessageBox.warning(self, "Добавление изображения",
                                            "Данный файл уже загружен", QMessageBox.Ok)

                    self.left_button_birth_certificate.setHidden(False)
                    self.right_button_birth_certificate.setHidden(False)
                    self.delete_image_birth_certificate.setHidden(False)

                elif self.sender() == self.open_explorer_snils:
                    if file not in self.snils_images:
                        self.snils_image.setPixmap(QPixmap(file))
                        self.snils_images.append(file)

                    else:
                        QMessageBox.warning(self, "Добавление изображения",
                                            "Данный файл уже загружен", QMessageBox.Ok)

                    self.left_button_snils.setHidden(False)
                    self.right_button_snils.setHidden(False)
                    self.delete_image_snils.setHidden(False)

                elif self.sender() == self.open_explorer_certificate:
                    if file not in self.certificate_images:
                        self.certificate_image.setPixmap(QPixmap(file))
                        self.certificate_images.append(file)

                    else:
                        QMessageBox.warning(self, "Добавление изображения",
                                            "Данный файл уже загружен", QMessageBox.Ok)

                    self.left_button_certificate.setHidden(False)
                    self.right_button_certificate.setHidden(False)
                    self.delete_image_certificate.setHidden(False)

                elif self.sender() == self.open_explorer_passport_2:
                    self.passport_image_2.setPixmap(QPixmap(file))

        except ValueError:
            QMessageBox.warning(self, "Ошибка при открытии", "Файл не является изображением или не поддерживается",
                                QMessageBox.Ok)

        except FileNotFoundError:
            pass

    def change_date(self):
        if self.sender() == self.calendar_birthday:
            self.date_of_birth.setDate(self.calendar_birthday.selectedDate())

        else:
            self.date_of_issue.setDate(self.calendar_issue.selectedDate())

    def move_attachments_certificate(self):
        self.certificate_image.setHidden(not self.certificate_image.isHidden())

        self.left_button_certificate.setHidden(self.certificate_image.isHidden())
        self.right_button_certificate.setHidden(self.certificate_image.isHidden())
        self.delete_image_certificate.setHidden(self.certificate_image.isHidden())

        if len(self.certificate_images) == 0:
            self.left_button_certificate.setHidden(True)
            self.right_button_certificate.setHidden(True)
            self.delete_image_certificate.setHidden(True)

    def move_widgets_in_first_column(self):
        self.passport_image.setHidden(not self.passport_image.isHidden())

        self.left_button_passport.setHidden(self.passport_image.isHidden())
        self.right_button_passport.setHidden(self.passport_image.isHidden())
        self.delete_image_passport.setHidden(self.passport_image.isHidden())

        if len(self.passport_images) == 0:
            self.left_button_passport.setHidden(True)
            self.right_button_passport.setHidden(True)
            self.delete_image_passport.setHidden(True)

        if not self.passport_image.isHidden():
            self.open_attachments_passport.setIcon(QIcon("images/delete.png"))

            self.label_14.move(self.label_14.geometry().x(),
                               self.label_14.geometry().y() + self.passport_image.size().height())

            self.label_15.move(self.label_15.geometry().x(),
                               self.label_15.geometry().y() + self.passport_image.size().height())

            self.label_18.move(self.label_18.geometry().x(),
                               self.label_18.geometry().y() + self.passport_image.size().height())

            self.issued_by_whom.move(self.issued_by_whom.geometry().x(),
                                     self.issued_by_whom.geometry().y() + self.passport_image.size().height())

            self.issue_error.move(self.issue_error.geometry().x(),
                                  self.issue_error.geometry().y() + self.passport_image.size().height())

            self.date_of_issue.move(self.date_of_issue.geometry().x(),
                                    self.date_of_issue.geometry().y() + self.passport_image.size().height())

            self.open_calendar_issue.move(self.open_calendar_issue.geometry().x(),
                                          self.open_calendar_issue.geometry().y() + self.passport_image.size().height())

            self.calendar_issue.move(self.calendar_issue.geometry().x(),
                                     self.calendar_issue.geometry().y() + self.passport_image.size().height())

            self.date_of_issue_error.move(self.date_of_issue_error.geometry().x(),
                                          self.date_of_issue_error.geometry().y() + self.passport_image.size().height())

            self.division_code.move(self.division_code.geometry().x(),
                                    self.division_code.geometry().y() + self.passport_image.size().height())

            self.division_code_error.move(self.division_code_error.geometry().x(),
                                          self.division_code_error.geometry().y() + self.passport_image.size().height())

        else:
            self.open_attachments_passport.setIcon(QIcon("images/attachments.png"))

            self.label_14.move(self.label_14.geometry().x(),
                               self.label_14.geometry().y() - self.passport_image.size().height())

            self.label_15.move(self.label_15.geometry().x(),
                               self.label_15.geometry().y() - self.passport_image.size().height())

            self.label_18.move(self.label_18.geometry().x(),
                               self.label_18.geometry().y() - self.passport_image.size().height())

            self.issued_by_whom.move(self.issued_by_whom.geometry().x(),
                                     self.issued_by_whom.geometry().y() - self.passport_image.size().height())

            self.issue_error.move(self.issue_error.geometry().x(),
                                  self.issue_error.geometry().y() - self.passport_image.size().height())

            self.date_of_issue.move(self.date_of_issue.geometry().x(),
                                    self.date_of_issue.geometry().y() - self.passport_image.size().height())

            self.open_calendar_issue.move(self.open_calendar_issue.geometry().x(),
                                          self.open_calendar_issue.geometry().y() - self.passport_image.size().height())

            self.calendar_issue.move(self.calendar_issue.geometry().x(),
                                     self.calendar_issue.geometry().y() - self.passport_image.size().height())

            self.date_of_issue_error.move(self.date_of_issue_error.geometry().x(),
                                          self.date_of_issue_error.geometry().y() - self.passport_image.size().height())

            self.division_code.move(self.division_code.geometry().x(),
                                    self.division_code.geometry().y() - self.passport_image.size().height())

            self.division_code_error.move(self.division_code_error.geometry().x(),
                                          self.division_code_error.geometry().y() - self.passport_image.size().height())

    def move_widgets_in_second_column(self):
        self.snils_image.setHidden(not self.snils_image.isHidden())

        self.left_button_snils.setHidden(self.snils_image.isHidden())
        self.right_button_snils.setHidden(self.snils_image.isHidden())
        self.delete_image_snils.setHidden(self.snils_image.isHidden())

        if len(self.snils_images) == 0:
            self.left_button_snils.setHidden(True)
            self.right_button_snils.setHidden(True)
            self.delete_image_snils.setHidden(True)

        if self.snils_image.isHidden():
            self.open_attachments_snils.setIcon(QIcon("images/attachments.png"))

            self.label_58.move(self.label_58.geometry().x(),
                               self.label_58.geometry().y() - self.snils_image.size().height())

            self.oms_policy_number.move(self.oms_policy_number.geometry().x(),
                                        self.oms_policy_number.geometry().y() - self.snils_image.size().height())

            self.open_attachments_oms.move(self.open_attachments_oms.geometry().x(),
                                           self.open_attachments_oms.geometry().y() - self.snils_image.size().height())

            self.open_explorer_oms.move(self.open_explorer_oms.geometry().x(),
                                        self.open_explorer_oms.geometry().y() - self.snils_image.size().height())

            self.oms_policy_number_error.move(self.oms_policy_number_error.geometry().x(),
                                              self.oms_policy_number_error.geometry().y() - self.snils_image.size().height())

            self.oms_image.move(self.oms_image.geometry().x(),
                                self.oms_image.geometry().y() - self.snils_image.size().height())

            self.left_button_oms.move(self.left_button_oms.geometry().x(),
                                      self.left_button_oms.geometry().y() - self.snils_image.size().height())

            self.right_button_oms.move(self.right_button_oms.geometry().x(),
                                       self.right_button_oms.geometry().y() - self.snils_image.size().height())

            self.delete_image_oms.move(self.delete_image_oms.geometry().x(),
                                       self.delete_image_oms.geometry().y() - self.snils_image.size().height())

            self.label_59.move(self.label_59.geometry().x(),
                               self.label_59.geometry().y() - self.snils_image.size().height())

            self.birth_certificate_number.move(self.birth_certificate_number.geometry().x(),
                                               self.birth_certificate_number.geometry().y() - self.snils_image.size().height())

            self.open_attachments_birth_certificate.move(self.open_attachments_birth_certificate.geometry().x(),
                                                         self.open_attachments_birth_certificate.geometry().y() - self.snils_image.size().height())

            self.open_explorer_birth_certificate.move(self.open_explorer_birth_certificate.geometry().x(),
                                                      self.open_explorer_birth_certificate.geometry().y() - self.snils_image.size().height())

            self.birth_certificate_number_error.move(self.birth_certificate_number_error.geometry().x(),
                                                     self.birth_certificate_number_error.geometry().y() - self.snils_image.size().height())

            self.birth_certificate_image.move(self.birth_certificate_image.geometry().x(),
                                              self.birth_certificate_image.geometry().y() - self.snils_image.size().height())

            self.left_button_birth_certificate.move(self.left_button_birth_certificate.geometry().x(),
                                                    self.left_button_birth_certificate.geometry().y() - self.snils_image.size().height())

            self.right_button_birth_certificate.move(self.right_button_birth_certificate.geometry().x(),
                                                     self.right_button_birth_certificate.geometry().y() - self.snils_image.size().height())

            self.delete_image_birth_certificate.move(self.delete_image_birth_certificate.geometry().x(),
                                                     self.delete_image_birth_certificate.geometry().y() - self.snils_image.size().height())

        else:
            self.open_attachments_snils.setIcon(QIcon("images/delete.png"))

            self.label_58.move(self.label_58.geometry().x(),
                               self.label_58.geometry().y() + self.snils_image.size().height())

            self.oms_policy_number.move(self.oms_policy_number.geometry().x(),
                                        self.oms_policy_number.geometry().y() + self.snils_image.size().height())

            self.open_attachments_oms.move(self.open_attachments_oms.geometry().x(),
                                           self.open_attachments_oms.geometry().y() + self.snils_image.size().height())

            self.open_explorer_oms.move(self.open_explorer_oms.geometry().x(),
                                        self.open_explorer_oms.geometry().y() + self.snils_image.size().height())

            self.oms_policy_number_error.move(self.oms_policy_number_error.geometry().x(),
                                              self.oms_policy_number_error.geometry().y() + self.snils_image.size().height())

            self.oms_image.move(self.oms_image.geometry().x(),
                                self.oms_image.geometry().y() + self.snils_image.size().height())

            self.left_button_oms.move(self.left_button_oms.geometry().x(),
                                      self.left_button_oms.geometry().y() + self.snils_image.size().height())

            self.right_button_oms.move(self.right_button_oms.geometry().x(),
                                       self.right_button_oms.geometry().y() + self.snils_image.size().height())

            self.delete_image_oms.move(self.delete_image_oms.geometry().x(),
                                       self.delete_image_oms.geometry().y() + self.snils_image.size().height())

            self.label_59.move(self.label_59.geometry().x(),
                               self.label_59.geometry().y() + self.snils_image.size().height())

            self.birth_certificate_number.move(self.birth_certificate_number.geometry().x(),
                                               self.birth_certificate_number.geometry().y() + self.snils_image.size().height())

            self.open_attachments_birth_certificate.move(self.open_attachments_birth_certificate.geometry().x(),
                                                         self.open_attachments_birth_certificate.geometry().y() + self.snils_image.size().height())

            self.open_explorer_birth_certificate.move(self.open_explorer_birth_certificate.geometry().x(),
                                                      self.open_explorer_birth_certificate.geometry().y() + self.snils_image.size().height())

            self.birth_certificate_number_error.move(self.birth_certificate_number_error.geometry().x(),
                                                     self.birth_certificate_number_error.geometry().y() + self.snils_image.size().height())

            self.birth_certificate_image.move(self.birth_certificate_image.geometry().x(),
                                              self.birth_certificate_image.geometry().y() + self.snils_image.size().height())

            self.left_button_birth_certificate.move(self.left_button_birth_certificate.geometry().x(),
                                                    self.left_button_birth_certificate.geometry().y() + self.snils_image.size().height())

            self.right_button_birth_certificate.move(self.right_button_birth_certificate.geometry().x(),
                                                     self.right_button_birth_certificate.geometry().y() + self.snils_image.size().height())

            self.delete_image_birth_certificate.move(self.delete_image_birth_certificate.geometry().x(),
                                                     self.delete_image_birth_certificate.geometry().y() + self.snils_image.size().height())

    def move_widgets_in_second_column_2(self):
        self.oms_image.setHidden(not self.oms_image.isHidden())

        self.left_button_oms.setHidden(self.oms_image.isHidden())
        self.right_button_oms.setHidden(self.oms_image.isHidden())
        self.delete_image_oms.setHidden(self.oms_image.isHidden())

        if len(self.oms_images) == 0:
            self.left_button_oms.setHidden(True)
            self.right_button_oms.setHidden(True)
            self.delete_image_oms.setHidden(True)

        if self.oms_image.isHidden():
            self.open_attachments_oms.setIcon(QIcon("images/attachments.png"))

            self.label_59.move(self.label_59.geometry().x(),
                               self.label_59.geometry().y() - self.oms_image.size().height())

            self.birth_certificate_number.move(self.birth_certificate_number.geometry().x(),
                                               self.birth_certificate_number.geometry().y() - self.oms_image.size().height())

            self.open_attachments_birth_certificate.move(self.open_attachments_birth_certificate.geometry().x(),
                                                         self.open_attachments_birth_certificate.geometry().y() - self.oms_image.size().height())

            self.open_explorer_birth_certificate.move(self.open_explorer_birth_certificate.geometry().x(),
                                                      self.open_explorer_birth_certificate.geometry().y() - self.oms_image.size().height())

            self.birth_certificate_number_error.move(self.birth_certificate_number_error.geometry().x(),
                                                     self.birth_certificate_number_error.geometry().y() - self.oms_image.size().height())

            self.birth_certificate_image.move(self.birth_certificate_image.geometry().x(),
                                              self.birth_certificate_image.geometry().y() - self.oms_image.size().height())

            self.left_button_birth_certificate.move(self.left_button_birth_certificate.geometry().x(),
                                                    self.left_button_birth_certificate.geometry().y() - self.oms_image.size().height())

            self.right_button_birth_certificate.move(self.right_button_birth_certificate.geometry().x(),
                                                     self.right_button_birth_certificate.geometry().y() - self.oms_image.size().height())

            self.delete_image_birth_certificate.move(self.delete_image_birth_certificate.geometry().x(),
                                                     self.delete_image_birth_certificate.geometry().y() - self.oms_image.size().height())

        else:
            self.open_attachments_oms.setIcon(QIcon("images/delete.png"))

            self.label_59.move(self.label_59.geometry().x(),
                               self.label_59.geometry().y() + self.oms_image.size().height())

            self.birth_certificate_number.move(self.birth_certificate_number.geometry().x(),
                                               self.birth_certificate_number.geometry().y() + self.oms_image.size().height())

            self.open_attachments_birth_certificate.move(self.open_attachments_birth_certificate.geometry().x(),
                                                         self.open_attachments_birth_certificate.geometry().y() + self.oms_image.size().height())

            self.open_explorer_birth_certificate.move(self.open_explorer_birth_certificate.geometry().x(),
                                                      self.open_explorer_birth_certificate.geometry().y() + self.oms_image.size().height())

            self.birth_certificate_number_error.move(self.birth_certificate_number_error.geometry().x(),
                                                     self.birth_certificate_number_error.geometry().y() + self.oms_image.size().height())

            self.birth_certificate_image.move(self.birth_certificate_image.geometry().x(),
                                              self.birth_certificate_image.geometry().y() + self.oms_image.size().height())

            self.left_button_birth_certificate.move(self.left_button_birth_certificate.geometry().x(),
                                                    self.left_button_birth_certificate.geometry().y() + self.oms_image.size().height())

            self.right_button_birth_certificate.move(self.right_button_birth_certificate.geometry().x(),
                                                     self.right_button_birth_certificate.geometry().y() + self.oms_image.size().height())

            self.delete_image_birth_certificate.move(self.delete_image_birth_certificate.geometry().x(),
                                                     self.delete_image_birth_certificate.geometry().y() + self.oms_image.size().height())

    def move_widgets_in_second_column_3(self):
        self.birth_certificate_image.setHidden(not self.birth_certificate_image.isHidden())

        self.left_button_birth_certificate.setHidden(self.birth_certificate_image.isHidden())
        self.right_button_birth_certificate.setHidden(self.birth_certificate_image.isHidden())
        self.delete_image_birth_certificate.setHidden(self.birth_certificate_image.isHidden())

        if len(self.birth_certificate_images) == 0:
            self.left_button_birth_certificate.setHidden(True)
            self.right_button_birth_certificate.setHidden(True)
            self.delete_image_birth_certificate.setHidden(True)

        if self.birth_certificate_image.isHidden():
            self.open_attachments_birth_certificate.setIcon(QIcon("images/attachments.png"))

        else:
            self.open_attachments_birth_certificate.setIcon(QIcon("images/delete.png"))

    def open_and_close_calendar_issue(self):
        self.calendar_issue.setHidden(not self.calendar_issue.isHidden())

        if self.calendar_issue.isHidden():
            self.open_calendar_issue.setIcon(QIcon("images/calendar.png"))

            self.date_of_issue_error.move(self.date_of_issue_error.geometry().x(),
                                          self.date_of_issue_error.geometry().y() - self.calendar_issue.size().height())

            self.label_18.move(self.label_18.geometry().x(),
                               self.label_18.geometry().y() - self.calendar_issue.size().height())

            self.division_code.move(self.division_code.geometry().x(),
                                    self.division_code.geometry().y() - self.calendar_issue.size().height())

            self.division_code_error.move(self.division_code_error.geometry().x(),
                                          self.division_code_error.geometry().y() - self.calendar_issue.size().height())

        else:
            self.open_calendar_issue.setIcon(QIcon("images/delete.png"))

            self.date_of_issue_error.move(self.date_of_issue_error.geometry().x(),
                                          self.date_of_issue_error.geometry().y() + self.calendar_issue.size().height())

            self.label_18.move(self.label_18.geometry().x(),
                               self.label_18.geometry().y() + self.calendar_issue.size().height())

            self.division_code.move(self.division_code.geometry().x(),
                                    self.division_code.geometry().y() + self.calendar_issue.size().height())

            self.division_code_error.move(self.division_code_error.geometry().x(),
                                          self.division_code_error.geometry().y() + self.calendar_issue.size().height())

    def open_and_hide_calendar_birthday(self):
        if self.flag:
            self.calendar_birthday.show()
            self.open_calendar_birthday.setIcon(QIcon("images/delete.png"))

        if not self.flag:
            self.calendar_birthday.hide()
            self.open_calendar_birthday.setIcon(QIcon("images/calendar.png"))

        self.flag = not self.flag

    def length_text_edit(self):
        if len(self.about_applicant_textedit.toPlainText()) >= 999:
            self.about_applicant_textedit.setPlainText(self.about_applicant_textedit.toPlainText()[:999])
        self.length_textedit.setText(f"{len(self.about_applicant_textedit.toPlainText())}/999")

    def close_and_open_side_panel(self):
        if self.flag2:
            self.open_and_close_button.setIcon(QIcon("images/three_lines.png"))

            self.anim2 = QPropertyAnimation(self.horizontalFrame, b"size")
            self.anim2.setStartValue(self.horizontalFrame.size())
            self.anim2.setEndValue(QSize(53, self.horizontalFrame.size().height()))
            self.anim2.setDuration(150)
            self.anim2.start()

        else:
            self.anim2 = QPropertyAnimation(self.horizontalFrame, b"size")
            self.anim2.setStartValue(self.horizontalFrame.size())
            self.anim2.setEndValue(QSize(259, self.horizontalFrame.size().height()))
            self.anim2.setDuration(150)
            self.anim2.start()

            self.open_and_close_button.setIcon(QIcon("images/delete.png"))

        if self.flag2:
            self.verticalFrame_2.hide()

        else:
            self.verticalFrame_2.show()

        self.flag2 = not self.flag2

    def change_cities_and_villages(self):
        if self.region.currentText() == "Республика Дагестан":
            self.city.clear()
            self.city.addItems(self.dagestan_cities_and_villages)

        elif self.region.currentText() == "Чеченская Республика":
            self.city.clear()
            self.city.addItems(self.chechnya_cities_and_villages)

    def erase_the_inscription(self):
        if self.sender() == self.fio:
            self.fio_error.setText("")

        elif self.sender() == self.date_of_birth:
            self.age_student_error.setText("")

        elif self.sender() == self.address:
            self.address_error.setText("")

        elif self.sender() == self.series_and_number_passport:
            self.passport_error.setText("")

        elif self.sender() == self.issued_by_whom:
            self.issue_error.setText("")

        elif self.sender() == self.date_of_issue:
            self.date_of_issue_error.setText("")

        elif self.sender() == self.snils:
            self.snils_error.setText("")

        elif self.sender() == self.certificate_number:
            self.certificate_number_error.setText("")

    def check_data(self):
        error = False
        # проверка фио
        for i in self.fio.text():
            if not i.isalpha() and i != " ":
                self.fio_error.setText("В поле присутствуют недопустимые символы")
                error = True
                break

        if len(self.fio.text()) == 0:
            error = True
            self.fio_error.setText("Введите ФИО")

        # проверка даты рождения
        currently_date = list(map(int, str(datetime.datetime.now())[:10].split("-")))

        birth = list(map(int, self.date_of_birth.text().split(".")))
        birth_in_currently_year = datetime.date(currently_date[0], birth[1], birth[0])

        currently_date = datetime.date(currently_date[0], currently_date[1], currently_date[2])
        birth = datetime.date(birth[2], birth[1], birth[0])

        if str(currently_date - birth) == "0:00:00":
            error = True
            self.age_student_error.setText("Он не мог родиться сегодня")

        else:
            if int(str(currently_date - birth).split()[0]) < 0:
                error = True
                self.age_student_error.setText("Абитуриент из будущего?")

            elif int(str(currently_date - birth).split()[0]) // 365 < 16:
                error = True
                self.age_student_error.setText("Возраст абитуриента меньше 16-ти лет")

            elif int(str(birth_in_currently_year - birth).split()[0]) // 365 > 30:
                error = True
                self.age_student_error.setText("Возраст абитуриента превышает 30 лет")

        # проверка адреса
        for i in self.address.text():
            if not i.isalnum() and i not in '"№' and i not in " ,.":
                self.address_error.setText("В адресе содержатся недопустимые символы")
                error = True
                break

        if len(self.address.text()) == 0:
            self.address_error.setText("Введите адрес")
            error = True

        # проверка номера аттестата
        if len(self.certificate_number.text()) != 14:
            self.certificate_number_error.setText("В номере должно быть 14 чисел")
            error = True

        for i in self.certificate_number.text():
            if not i.isdigit():
                self.certificate_number_error.setText("В поле не может быть других знаков, кроме чисел")
                error = True
                break

        # проверка номера и серии паспорта
        digits_count = 0

        for i in self.series_and_number_passport.text():
            if i.isdigit():
                digits_count += 1

        if len(self.series_and_number_passport.text()) != 11 and digits_count != 10:
            error = True
            self.passport_error.setText("В поле должно быть 10 чисел, не считая пробела")

        for i in self.series_and_number_passport.text():
            if not i.isdigit() and i not in " ":
                error = True
                self.passport_error.setText("В поле не может быть других знаков, кроме чисел")
                break

        # проверка поля "Кем выдан?"
        for i in self.issued_by_whom.text():
            if not i.isalpha() and i != " ":
                error = True
                self.issue_error.setText("В поле не может быть ничего, кроме букв")

        if len(self.issued_by_whom.text()) == 0:
            error = True
            self.issue_error.setText("Заполните данное поле")

        # проверка даты выдачи
        birth = list(map(int, self.date_of_birth.text().split(".")))
        birth = datetime.date(birth[2], birth[1], birth[0])

        issue = list(map(int, self.date_of_issue.text().split(".")))
        issue = datetime.date(issue[2], issue[1], issue[0])

        if str(issue - birth).split()[0] != "0:00:00":
            different = int(str(issue - birth).split()[0]) // 365
        else:
            different = "0"

        correct_word = "лет"
        if str(different)[-1] == "1" and len(str(different)) != 2:
            correct_word = "год"

        elif str(different)[-1] in "234" and len(str(different)) == 1:
            correct_word = "года"

        if different == "0":
            error = True
            self.date_of_issue_error.setText("Он не мог получить паспорт в день его рождения")

        elif different < 14:
            error = True
            self.date_of_issue_error.setText(f"На момент выдачи ему не могло быть {different} {correct_word}")

        # проверка номера аттестата
        if len(self.certificate_number.text()) != 14:
            error = True
            self.certificate_number_error.setText("Длина поля должна составлять 14 цифр")

        for i in self.certificate_number.text():
            if not i.isdigit():
                error = True
                self.certificate_number_error.setText("В поле не должно быть других знаков, кроме чисел")
                break

        # проверка снилса
        count_digits = 0

        for i in range(len(self.snils.text())):
            if self.snils.text()[i] in "0123456789":
                count_digits += 1

            if self.snils.text()[i] == "-":
                if count_digits != 3:
                    self.snils_error.setText('Введите снилс в формате: "nnn-nnn-nnn nn"')
                    error = True
                    break
                else:
                    count_digits = 0

            elif self.snils.text()[i] == " ":
                if count_digits != 3:
                    error = True
                    self.snils_error.setText('Введите снилс в формате: "nnn-nnn-nnn nn"')
                    break

        for i in self.snils.text():
            if not i.isdigit() and i != "-" and i != " ":
                error = True
                self.snils_error.setText('Введите снилс в формате: "nnn-nnn-nnn nn"')

        if len(self.snils.text()) != 14:
            error = True
            self.snils_error.setText('Введите снилс в формате: "nnn-nnn-nnn nn"')

        # проверка пути
        try:
            file = open(self.path.text())
            file.close()

        except FileNotFoundError:
            error = True

        if len(self.applicants_number.text()) == 0:
            self.applicants_number_error.setText("Введите номер телефона")
            error = True

        else:
            for latter in range(len(self.applicants_number.text())):
                if self.applicants_number.text()[latter].isalpha() and self.applicants_number.text()[latter] != "+":
                    self.applicants_number_error.setText("В номере не может быть чего-либо, кроме цифр и '+'")
                    error = True
                    break

        if not error:
            self.write_in_file()

    def write_in_file(self):
        if self.path.text().split(".")[-1] == "csv":
            dialog = Csv_encoding()
            dialog.show()

            if dialog.exec_():
                if dialog.encoding.currentText() == "UTF-8 (по умолчанию)":
                    self.encode = "utf8"

                elif dialog.encoding.currentText() == "ASCII":
                    self.encode = "ASCII"

                elif dialog.encoding.currentText() == "Windows-1251 (для кириллических алфавитов)":
                    self.encode = "Windows-1251"

            if self.encode != "":
                with open(self.path.text(), mode="a", encoding=self.encode) as file:
                    if dialog.separator.currentText() == "Точка с запятой":
                        self.separator = ";"
                        writer = csv.writer(file, delimiter=";")  # открываю файл

                    elif dialog.separator.currentText() == "Табуляция":
                        self.separator = "\t"
                        writer = csv.writer(file, delimiter="\t")

                    elif dialog.separator.currentText() == "Пробел":
                        self.separator = " "
                        writer = csv.writer(file, delimiter=" ")

                    elif dialog.separator.currentText() == "Запятая":
                        self.separator = ","
                        writer = csv.writer(file, delimiter=",")

                    elif dialog.separator.currentText() == "Двоеточие":
                        self.separator = ":"
                        writer = csv.writer(file, delimiter=":")

                    student = [self.fio.text(), self.date_of_birth.text(), self.sex.currentText(),
                               self.region.currentText(), self.city.currentText(),
                               self.series_and_number_passport.text(), self.issued_by_whom.text(),
                               self.date_of_issue.text(), self.snils.text(), self.certificate_number.text()]

                    writer.writerow(student)

                question = QMessageBox.question(self, "Добавление абитуриента", "Добавить ещё одного абитуриента?",
                                                QMessageBox.Yes, QMessageBox.No)

                if question == QMessageBox.Yes:
                    self.fio.setText("")
                    date_now = list(map(int, str(datetime.datetime.now())[:10].split("-")))

                    self.date_of_birth.setDate(QDate(date_now[0], date_now[1], date_now[2]))
                    self.date_of_issue.setDate(QDate(date_now[0], date_now[1], date_now[2]))

                    self.address.setText("")
                    self.path.setText("")
                    self.series_and_number_passport.setText("")
                    self.issued_by_whom.setText("")
                    self.snils.setText("")
                    self.certificate_number.setText("")

                else:
                    self.hide_dialog = True
                    main_w = Main_window()
                    main_w.show()
                    self.close()

    def set_date(self):
        if self.signal == 0:
            self.date_of_birth.setDate(self.calendar.selectedDate())

        else:
            self.date_of_issue.setDate(self.calendar.selectedDate())

    def confirm_exit(self):
        self.close()

    def get_path(self):
        get_path = \
            QFileDialog.getOpenFileName(self, "Выберите файл", "/",
                                        "Таблица (*.xls, *.xlsx);; CSV-таблица(*.csv);;База данных(*.sqlite)")[0]

        self.path.setText(get_path)


class Viewing_applicants(Ui_viewing_applicants, QMainWindow):
    def __init__(self):
        super(Ui_viewing_applicants, self).__init__()
        self.setupUi(self)
        self.process()

    def process(self):
        self.pushButton.clicked.connect(self.return_to_main_menu)

    def return_to_main_menu(self):
        self.window = Main_window()
        self.window.show()
        self.close()


# окно печати документов
class Print_documents(Ui_print, QMainWindow):
    def __init__(self):
        super(Ui_print, self).__init__()
        self.setupUi(self)
        self.process()

    def process(self):
        self.pushButton_4.clicked.connect(self.exit_to_main_menu)

    def exit_to_main_menu(self):
        self.main = Main_window()
        self.main.show()
        self.close()


# окно "О программе"
class About(Ui_about_programm, QMainWindow):
    def __init__(self):
        super(Ui_about_programm, self).__init__()
        self.setupUi(self)
        self.setStyleSheet("background-color:rgb(255, 255, 255)")
        self.setWindowIcon(QIcon("images/icon.png"))
        pixmap = QPixmap("images/icon.png")

        self.picture.setPixmap(pixmap)


if __name__ == '__main__':
    app = QApplication(sys.argv)
    ex = Main_window()
    ex.show()
    sys.excepthook = except_hook
    sys.exit(app.exec())
